from codecs import BOM_UTF8
import os
import re
import tempfile
import zipfile
import csv
import json
from django.template import Context
from django.template.loader import render_to_string, get_template
import xlwt


class UniqueHeaderGenerator(object):
    def __init__(self, max_column_size=None):
        self.used = set()
        self.max_column_size = max_column_size or 2000

    def next_unique(self, header):
        header = self._next_unique(header)
        self.used.add(header)
        return header

    def _next_unique(self, string):
        counter = 1
        if len(string) > self.max_column_size:
            # truncate from the beginning since the end has more specific information
            string = string[-self.max_column_size:]
        orig_string = string
        while string in self.used:
            string = "%s%s" % (orig_string, counter)
            if len(string) > self.max_column_size:
                counterlen = len(str(counter))
                string = "%s%s" % (orig_string[-(self.max_column_size - counterlen):], counter)
            counter += 1

        return string

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass


class ExportFileWriter(object):

    def __init__(self):
        self.name = None
        self._isopen = False
        self._file = None
        self._path = None

    def get_path(self):
        assert self._isopen
        return self._path

    def get_file(self):
        assert self._isopen
        return self._file

    def open(self, name):
        assert not self._isopen
        self._isopen = True
        self.name = name
        fd, path = tempfile.mkstemp()
        self._file = os.fdopen(fd, 'wb+')
        self._path = path
        self._open()
        self._begin_file()

    def _open(self):
        pass

    def _begin_file(self):
        pass

    def write_row(self, row):
        raise NotImplementedError

    def _end_file(self):
        pass

    def finish(self):
        self._end_file()
        self._file.seek(0)

    def close(self):
        assert self._isopen
        self._file.close()
        os.remove(self._path)
        self._isopen = False


class CsvFileWriter(ExportFileWriter):

    def _open(self):
        # Excel needs UTF8-encoded CSVs to start with the UTF-8 byte-order mark (FB 163268)
        self._file.write(BOM_UTF8)
        self._csvwriter = csv.writer(self._file, csv.excel)

    def write_row(self, row):
        self._csvwriter.writerow(row)


class PartialHtmlFileWriter(ExportFileWriter):

    def _write_from_template(self, context):
        self._file.write(self.template.render(Context(context)).encode('utf-8'))

    def _open(self):
        self.template = get_template("couchexport/html_export.html")
        self._on_first_row = True

    def write_row(self, row):
        section = "row" if not self._on_first_row else "first_row"
        self._on_first_row = False
        self._write_from_template({"row": row, "section": section})

    def _end_file(self):
        if self._on_first_row:
            # There were no rows
            self._write_from_template({"section": "no_rows"})


class HtmlFileWriter(PartialHtmlFileWriter):

    def _begin_file(self):
        self._write_from_template({"section": "doc_begin"})
        self._write_from_template({"section": "table_begin", "name": self.name})

    def _end_file(self):
        super(HtmlFileWriter, self)._end_file()
        self._write_from_template({"section": "table_end"})
        self._write_from_template({"section": "doc_end"})


class ExportWriter(object):
    max_table_name_size = 500

    def open(self, header_table, file, max_column_size=2000, table_titles=None):
        """
        Create any initial files, headings, etc necessary.
        """
        table_titles = table_titles or {}

        self._isopen = True
        self.max_column_size = max_column_size
        self._current_primary_id = 0
        self.file = file

        self._init()
        self.table_name_generator = UniqueHeaderGenerator(
            self.max_table_name_size
        )
        for table_index, table in header_table:
            self.add_table(
                table_index,
                table[0],
                table_title=table_titles.get(table_index)
            )

    def add_table(self, table_index, headers, table_title=None):
        def _clean_name(name):
            return re.sub(r"[\n]", '', re.sub(r"[[\\?*/:\]]", "-", unicode(name)))

        table_title_truncated = _clean_name(
            self.table_name_generator.next_unique(table_title or table_index)
        )

        # make sure we trim the headers
        with UniqueHeaderGenerator(self.max_column_size) as g:
            try:
                headers.data = [g.next_unique(header) for header in headers.data]
            except AttributeError:
                headers = [g.next_unique(header) for header in headers]

        self._init_table(table_index, table_title_truncated)
        self.write_row(table_index, headers)

    def write(self, document_table, skip_first=False):
        """
        Given a document that's been parsed into the appropriate
        rows, write those rows to the resulting files.
        """
        assert self._isopen
        for table_index, table in document_table:
            for i, row in enumerate(table):
                if skip_first and i is 0:
                    continue
                # update the primary component of the ID to match
                # how many docs we've seen
                try:
                    row_has_id = row.has_id()
                except AttributeError:
                    row_has_id = False
                if row_has_id:
                    row.id = (self._current_primary_id,) + tuple(row.id[1:])

                self.write_row(table_index, row)

        self._current_primary_id += 1

    def write_row(self, table_index, headers):
        """
        Currently just calls the subclass's implementation
        but if we were to add a universal validation step,
        such a thing would happen here.
        """
        return self._write_row(table_index, headers)

    def close(self):
        """
        Close any open file references, do any cleanup.
        """
        assert(self._isopen)
        self._close()
        self._isopen = False

    def _init(self):
        raise NotImplementedError

    def _init_table(self, table_index, table_title):
        raise NotImplementedError

    def _write_row(self, sheet_index, row):
        raise NotImplementedError

    def _close(self):
        raise NotImplementedError

    @classmethod
    def get_data(cls, row):
        """
        Get around the fact that row can be either an iterable or
        a FormattedRow

        """
        try:
            return row.get_data()
        except AttributeError:
            return row


class OnDiskExportWriter(ExportWriter):
    """
    Keeps tables in temporary csv files. Subclassed by other export writers.
    """
    writer_class = CsvFileWriter

    def _init(self):
        self.tables = {}
        self.table_names = {}

    def _init_table(self, table_index, table_title):
        writer = self.writer_class()
        self.tables[table_index] = writer
        writer.open(table_title)
        self.table_names[table_index] = table_title

    def _write_row(self, sheet_index, row):

        def _encode_if_needed(val):
            return val.encode("utf8") if isinstance(val, unicode) else val
        row = map(_encode_if_needed, self.get_data(row))

        self.tables[sheet_index].write_row(row)

    def _close(self):
        """
        Close any open file references, do any cleanup.
        """
        for writer in self.tables.values():
            writer.finish()

        self._write_final_result()

        for writer in self.tables.values():
            writer.close()

    def _write_final_result(self):
        """
        Subclasses should call this method then write to a zip file, html files, or whatever.
        """
        raise NotImplementedError


class ZippedExportWriter(OnDiskExportWriter):
    """
    Writer that creates a zip file containing a csv for each table.
    """
    table_file_extension = ".csv"

    def _write_final_result(self):

        archive = zipfile.ZipFile(self.file, 'w', zipfile.ZIP_DEFLATED)
        for index, name in self.table_names.items():
            if isinstance(name, unicode):
                name = name.encode('utf-8')
            path = self.tables[index].get_path()
            archive.write(path, "{}{}".format(name, self.table_file_extension))
        archive.close()
        self.file.seek(0)


class CsvExportWriter(ZippedExportWriter):
    """
    CSV writer that creates a zip file containing a csv for each table.
    """
    pass


class UnzippedCsvExportWriter(OnDiskExportWriter):
    """
    Serve the first table as a csv
    """

    def _write_final_result(self):

        tablefile = self.tables.values()[0].get_file()
        for line in tablefile:
            self.file.write(line)
        self.file.seek(0)


class Excel2007ExportWriter(ExportWriter):
    max_table_name_size = 31

    def _init(self):
        try:
            import openpyxl
        except ImportError:
            raise Exception("It doesn't look like this machine is configured for "
                            "excel export. To export to excel you have to run the "
                            "command:  easy_install openpyxl")

        self.book = openpyxl.Workbook(optimized_write=True)
        self.tables = {}
        self.table_indices = {}


    def _init_table(self, table_index, table_title):
        sheet = self.book.create_sheet()
        sheet.title = table_title
        self.tables[table_index] = sheet
        self.table_indices[table_index] = 0


    def _write_row(self, sheet_index, row):
        sheet = self.tables[sheet_index]

        # Source: http://stackoverflow.com/questions/1707890/fast-way-to-filter-illegal-xml-unicode-chars-in-python
        dirty_chars = re.compile(
            u'[\x00-\x08\x0b-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff\ufdd0-\ufddf\ufffe-\uffff]'
        )

        def get_write_value(value):
            if isinstance(value, (int, long, float)):
                return value
            if isinstance(value, str):
                value = unicode(value, encoding="utf-8")
            elif value is not None:
                value = unicode(value)
            else:
                value = u''
            return dirty_chars.sub(u'?', value)

        # NOTE: don't touch this. changing anything like formatting in the
        # row by referencing the cells will cause huge memory issues.
        # see: http://openpyxl.readthedocs.org/en/latest/optimized.html
        sheet.append(map(get_write_value, self.get_data(row)))

    def _close(self):
        """
        Close any open file references, do any cleanup.
        """
        self.book.save(self.file)


class Excel2003ExportWriter(ExportWriter):
    max_table_name_size = 31

    def _init(self):
        self.book = xlwt.Workbook()
        self.tables = {}
        self.table_indices = {}

    def _init_table(self, table_index, table_title):
        sheet = self.book.add_sheet(table_title)
        self.tables[table_index] = sheet
        self.table_indices[table_index] = 0

    def _write_row(self, sheet_index, row):
        row_index = self.table_indices[sheet_index]
        sheet = self.tables[sheet_index]
        # have to deal with primary ids
        for i, val in enumerate(self.get_data(row)):
            sheet.write(row_index,i,unicode(val))
        self.table_indices[sheet_index] = row_index + 1

    def _close(self):
        self.book.save(self.file)

class InMemoryExportWriter(ExportWriter):
    """
    Keeps tables in memory. Subclassed by other export writers.
    """

    def _init(self):
        self.tables = {}
        self.table_names = {}

    def _init_table(self, table_index, table_title):
        self.table_names[table_index] = table_title
        self.tables[table_index] = []

    def _write_row(self, sheet_index, row):
        table = self.tables[sheet_index]
        # have to deal with primary ids
        row_data = [val for val in self.get_data(row)]
        table.append(row_data)

    def _close(self):
        pass

class JsonExportWriter(InMemoryExportWriter):
    """
    Write tables to JSON
    """

    class ConstantEncoder(json.JSONEncoder):

        def default(self, obj):
            from dimagi.utils.web import json_handler
            from couchexport.export import Constant
            if isinstance(obj, Constant):
                return obj.message
            else:
                return json_handler(obj)

    def _close(self):
        new_tables = {}
        for tablename, data in self.tables.items():
            new_tables[self.table_names[tablename]] = {"headers":data[0], "rows": data[1:]}

        self.file.write(json.dumps(new_tables, cls=self.ConstantEncoder))


class HtmlExportWriter(OnDiskExportWriter):
    """
    Write tables to a single HTML file.
    """
    writer_class = PartialHtmlFileWriter

    def _write_final_result(self):

        def write(context):
            self.file.write(
                render_to_string(
                    "couchexport/html_export.html", context
                ).encode("utf-8")
            )

        write({"section": "doc_begin"})
        for index, name in self.table_names.items():
            table_writer = self.tables[index]
            write({"section": "table_begin", "name": name})
            for line in table_writer.get_file():
                self.file.write(line)
            write({"section": "table_end"})
        write({"section": "doc_end"})

        self.file.seek(0)


class ZippedHtmlExportWriter(ZippedExportWriter):
    """
    Write each table to an HTML file in a zipfile
    """
    writer_class = HtmlFileWriter
    table_file_extension = ".html"

